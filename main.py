import pandas as pd
import openpyxl
import streamlit as st
import os

def main():
    # ダッシュボード設定
    st.set_page_config(
        page_title="Skillnoteデータ→Excel力量マップ",
        layout="wide"
    )

    # Streamlitアプリのタイトル
    st.title("Skillnoteデータ➡編集用Excel力量マップ")

    # HTMLとCSSを使ってフォントサイズを指定
    html_str = f"""
    <style>p.big-font {{font-size:24px;color:black}}</style>
    <p class='big-font'>
    competence_map_related_item_for_map.csvとdownload.xlsxファイルから<br>
    力量マップ加工用の「スキルマップデータ.xlsx」を作る
    </p>
    """
    # st.write()を使ってHTMLを表示
    st.write(html_str, unsafe_allow_html=True)

    # カスタムCSSを定義
    custom_css = """
    <style>
        /* ファイルアップローダーのラベルスタイルを変更 */
            div.stFileUploader > div > label {
            font-size: 24px; /* フォントサイズを変更 */
            font-weight: bold; /* フォントを太字に */
            color: blue; /* フォントの色を変更 */
        /* その他のCSSプロパティを追加可能 */
        }
        .css-1v3fv5i{
            width:50px;
        }
    </style>
    """
    # カスタムCSSをアプリに埋め込む
    st.markdown(custom_css, unsafe_allow_html=True)

    # ファイルアップローダーを作成
    uploaded_file_csv = st.file_uploader("competence_map_related_item_for_map.csvファイルを選択してください", type=['csv'])
    uploaded_file_xlsx = st.file_uploader("download.xlsxファイルを選択してください", type=['xlsx'])

    # 両方のファイルがアップロードされたら処理を実行
    if uploaded_file_csv and uploaded_file_xlsx:
        # 一つ目のCSVファイルを読み込む
        df_csv = pd.read_csv(uploaded_file_csv)

        # 各行の最後の有効なセルの値を配列に格納
        values_array = []
        for index, row in df_csv.iterrows():
            # dropna()でNaNを除外し、最後の値を取得
            last_valid_value = row.dropna().iloc[-1]
            values_array.append(last_valid_value)
            # プロジェクト名を読取り
            value_projectName = row[0]
            value_mapCode = row[1]
        # print(values_array)

        # 2つ目のExcelファイルを読み込む
        wb = openpyxl.load_workbook(uploaded_file_xlsx)
        ws = wb.active

        # 2つ目のExcelファイルの最後列に配列データを11行目から入れる
        for i, value in enumerate(values_array, start=11):  # 11行目から開始
            ws.cell(row=i, column=ws.max_column, value=value)
            ws.cell(row=1,column=9,value=value_projectName)
            ws.cell(row=2,column=9,value=value_mapCode)

        # 修正したExcelデータを保存をホームのdownloadディレクトリに格納
        # ユーザーのホームディレクトリを取得
        home_directory = os.path.expanduser('~')
        # 'download'ディレクトリのフルパスを作成
        download_directory = os.path.join(home_directory, 'Downloads')
        print(download_directory)
        output_filename = 'スキルマップデータ.xlsx'
        output_filepath = os.path.join(download_directory, output_filename)
        wb.save(output_filepath)

        # HTMLとCSSを使ってフォントサイズを指定
        html_str = """
        <style>p.big-font2 {font-size:24px;color:blue}</style>
        <p class='big-font2'>
        「スキルマップデータ.xlsx」をローカルのDownloadsフォルダーに格納しました。
        </p>
        """
        # st.write()を使ってHTMLを表示
        st.write(html_str, unsafe_allow_html=True)

        "本ツールを再度実行する際は、再読み込み（左上の右回転矢印）してください。"

if __name__ == '__main__':
    main()